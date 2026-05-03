import os, io
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file
from werkzeug.utils import secure_filename
from database import get_db
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)
app.secret_key = 'super_secret_siprak_key'

# Automatic database initialization and migration on startup
try:
    from database import init_db, migrate
    init_db()
    migrate()
except Exception as e:
    print(f"Failed to auto-migrate database: {e}")

GDRIVE_UPLOAD_FOLDER = r'G:\My Drive\Siprak'
LOCAL_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if os.path.isdir(r'G:\My Drive'):
    UPLOAD_FOLDER = GDRIVE_UPLOAD_FOLDER
else:
    UPLOAD_FOLDER = LOCAL_UPLOAD_FOLDER
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'zip'}
def allowed_file(fn):
    return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ======== HELPERS ========
def get_letter_grade(s):
    if s >= 85: return 'A'
    elif s >= 80: return 'B+'
    elif s >= 75: return 'B'
    elif s >= 70: return 'C+'
    elif s >= 60: return 'C'
    elif s >= 50: return 'D+'
    elif s >= 40: return 'D'
    else: return 'E'

def calculate_module_avg(g):
    tp = int(g.get('tp_score',0) or 0)
    prak = int(g.get('praktikum_score',0) or 0)
    modul = int(g.get('modul_score',0) or 0)
    return round((tp+prak+modul)/3, 1)

def calculate_total(grades_dict, modules, pembukuan_score=0):
    if not modules: return 0
    avgs = []
    for m in modules:
        g = grades_dict.get(m['id'], {})
        avgs.append(calculate_module_avg(g))
    pbk = int(pembukuan_score or 0)
    return round(min((sum(avgs) + pbk) / (len(avgs) + 1), 100), 1)

GRADE_LEGEND = [
    ('A', '≥ 85', 'var(--success)'),
    ('B+', '80–84', 'var(--success)'),
    ('B', '75–79', 'var(--success)'),
    ('C+', '70–74', 'var(--primary)'),
    ('C', '60–69', 'var(--primary)'),
    ('D+', '50–59', 'var(--danger)'),
    ('D', '40–49', 'var(--danger)'),
    ('E', '< 40', 'var(--danger)'),
]

def get_allowed_groups(asprak_name, course_id):
    conn = get_db()
    # All aspraks can see all groups in the course now since they input them manually
    groups = conn.execute('SELECT DISTINCT group_id FROM users WHERE role="PRAKTIKAN" AND course_id=? ORDER BY group_id', (course_id,)).fetchall()
    conn.close()
    return tuple(g['group_id'] for g in groups)

def is_admin_user(user_id):
    conn = get_db()
    u = conn.execute('SELECT is_admin FROM users WHERE id=?', (user_id,)).fetchone()
    conn.close()
    return u and u['is_admin'] == 1

@app.context_processor
def inject_user():
    user = None
    if 'user_id' in session:
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        conn.close()
    return dict(current_user=user)

@app.route('/')
def index():
    return render_template('index.html')

# ======== PRAKTIKAN ========
@app.route('/praktikan', methods=['GET'])
def praktikan_dashboard():
    conn = get_db()
    courses = conn.execute('SELECT * FROM courses').fetchall()
    sel_course = request.args.get('course_id', type=int)
    modules, groups, drive_link = [], [], None
    if sel_course:
        modules = conn.execute('SELECT * FROM modules WHERE course_id=?', (sel_course,)).fetchall()
        groups = conn.execute('SELECT DISTINCT group_id FROM users WHERE role="PRAKTIKAN" AND course_id=? ORDER BY group_id', (sel_course,)).fetchall()
        course_row = conn.execute('SELECT drive_link FROM courses WHERE id=?', (sel_course,)).fetchone()
        drive_link = course_row['drive_link'] if course_row else None
    conn.close()
    return render_template('praktikan.html', courses=courses, modules=modules, groups=groups, sel_course=sel_course, drive_link=drive_link)

@app.route('/praktikan/submit', methods=['POST'])
def praktikan_submit():
    group_id = request.form.get('group_id')
    module_id = request.form.get('module_id')
    praktikan_name = request.form.get('praktikan_name')
    if 'file' not in request.files:
        flash('Tidak ada file', 'error'); return redirect(url_for('praktikan_dashboard'))
    file = request.files['file']
    if file.filename == '':
        flash('Tidak ada file yang dipilih', 'error'); return redirect(url_for('praktikan_dashboard'))
    if file and allowed_file(file.filename):
        conn = get_db()
        module = conn.execute('SELECT is_open, deadline FROM modules WHERE id=?', (module_id,)).fetchone()
        if not module or not module['is_open']:
            flash('Pengumpulan untuk modul ini sudah ditutup!', 'error'); conn.close(); return redirect(url_for('praktikan_dashboard'))
        if module['deadline']:
            try:
                # Handle possible varying datetime formats. Standard datetime-local is YYYY-MM-DDTHH:MM
                deadline_dt = datetime.strptime(module['deadline'], '%Y-%m-%dT%H:%M')
                if datetime.now() > deadline_dt:
                    flash('Pengumpulan untuk modul ini sudah melewati batas waktu (deadline)!', 'error'); conn.close(); return redirect(url_for('praktikan_dashboard'))
            except ValueError:
                pass # If format fails, fallback to string comparison or just pass
        existing = conn.execute('SELECT * FROM submissions WHERE group_id=? AND module_id=?', (group_id, module_id)).fetchone()
        if existing:
            flash('Kelompok sudah mengumpulkan modul ini!', 'error'); conn.close(); return redirect(url_for('praktikan_dashboard'))
        user = conn.execute('SELECT id FROM users WHERE role="PRAKTIKAN" AND group_id=? AND name LIKE ?', (group_id, f'%{praktikan_name}%')).fetchone()
        uid = user['id'] if user else None
        fn = secure_filename(f"Kelompok_{group_id}_Modul_{module_id}_{file.filename}")
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
        conn.execute('INSERT INTO submissions (module_id, group_id, file_path, submitted_by) VALUES (?,?,?,?)', (module_id, group_id, fn, uid))
        conn.commit(); conn.close()
        flash('Modul berhasil dikumpulkan!', 'success')
    else:
        flash('Ekstensi file tidak diizinkan', 'error')
    return redirect(url_for('praktikan_dashboard'))

# ======== ASPRAK LOGIN ========
@app.route('/asprak/login', methods=['GET', 'POST'])
def asprak_login():
    if 'role' in session and session['role'] == 'ASPRAK':
        return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        password = request.form.get('password')
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE role="ASPRAK" AND LOWER(name)=LOWER(?) AND password=?', (name, password)).fetchone()
        conn.close()
        if user:
            session['user_id'] = user['id']
            session['role'] = user['role']
            return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
        else:
            flash('Login gagal, periksa nama atau password', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# ======== ASPRAK DASHBOARD ========
@app.route('/asprak')
def asprak_dashboard():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        session.clear()
        return redirect(url_for('asprak_login'))
    asprak_name = user['name']
    admin = user['is_admin'] == 1
    if admin:
        courses = conn.execute('SELECT * FROM courses').fetchall()
    else:
        courses = conn.execute('''
            SELECT c.* FROM courses c
            JOIN users u ON u.course_id = c.id
            WHERE u.name = ? AND u.role = 'ASPRAK'
        ''', (asprak_name,)).fetchall()
    # Course selection
    sel_course = request.args.get('course_id', type=int)
    if not sel_course:
        sel_course = user['course_id'] or (courses[0]['id'] if courses else None)
    
    co_asprak = False
    if not admin and sel_course:
        user_for_course = conn.execute('SELECT is_co_asprak FROM users WHERE name=? AND role="ASPRAK" AND course_id=?', (asprak_name, sel_course)).fetchone()
        if user_for_course:
            co_asprak = user_for_course['is_co_asprak'] == 1
    
    # Tentukan tab default
    active_tab = request.args.get('tab')
    if not active_tab:
        if admin:
            active_tab = 'gdrive'
        else:
            active_tab = 'grading'

    if not sel_course:
        conn.close()
        return render_template('asprak.html', modules=[], submissions=[], praktikans=[], courses=courses,
                               sel_course=None, admin=admin, aspraks=[], calculate_module_avg=calculate_module_avg,
                               grade_legend=GRADE_LEGEND, all_groups=[], is_co_asprak=co_asprak, course_drive_link=None, active_tab=active_tab)
    modules = conn.execute('SELECT * FROM modules WHERE course_id=?', (sel_course,)).fetchall()
    
    # Get drive link for selected course early so it's available even if no praktikans exist
    course_row = conn.execute('SELECT drive_link FROM courses WHERE id=?', (sel_course,)).fetchone()
    course_drive_link = course_row['drive_link'] if course_row else None
    
    allowed = get_allowed_groups(asprak_name, sel_course)
    if not allowed:
        conn.close()
        aspraks = []
        if admin:
            conn2 = get_db()
            aspraks = conn2.execute('SELECT * FROM users WHERE role="ASPRAK" AND course_id=?', (sel_course,)).fetchall()
            conn2.close()
        return render_template('asprak.html', modules=modules, submissions=[], praktikans=[], courses=courses,
                               sel_course=sel_course, admin=admin, aspraks=aspraks, calculate_module_avg=calculate_module_avg,
                               grade_legend=GRADE_LEGEND, all_groups=[], is_co_asprak=co_asprak, course_drive_link=course_drive_link, active_tab=active_tab)
    ph = ','.join('?' for _ in allowed)
    subs_raw = conn.execute(f'''SELECT s.*, m.name as module_name, u.name as submitter_name
        FROM submissions s JOIN modules m ON s.module_id=m.id LEFT JOIN users u ON s.submitted_by=u.id
        WHERE s.group_id IN ({ph}) AND m.course_id=? ORDER BY s.timestamp DESC''', (*allowed, sel_course)).fetchall()
    submissions = []
    for sub in subs_raw:
        fp = os.path.join(app.config['UPLOAD_FOLDER'], sub['file_path'])
        if os.path.exists(fp):
            submissions.append(dict(sub))
        else:
            conn.execute('DELETE FROM submissions WHERE id=?', (sub['id'],)); conn.commit()
    praks_raw = conn.execute(f'SELECT * FROM users WHERE role="PRAKTIKAN" AND course_id=? AND group_id IN ({ph}) ORDER BY group_id, name',
                             (sel_course, *allowed)).fetchall()
    ml = [dict(m) for m in modules]
    praktikans = []
    for p in praks_raw:
        pd = dict(p)
        gr = conn.execute('SELECT * FROM grades WHERE praktikan_id=?', (p['id'],)).fetchall()
        pd['grades'] = {g['module_id']: dict(g) for g in gr}
        pd['total'] = calculate_total(pd['grades'], ml, pd.get('pembukuan_score', 0))
        pd['letter'] = get_letter_grade(pd['total'])
        praktikans.append(pd)
    aspraks = []
    if admin:
        aspraks = conn.execute('SELECT * FROM users WHERE role="ASPRAK" AND course_id=? ORDER BY name', (sel_course,)).fetchall()
    all_groups = conn.execute('SELECT DISTINCT group_id FROM users WHERE role="PRAKTIKAN" AND course_id=? ORDER BY group_id', (sel_course,)).fetchall()
    all_groups = [g['group_id'] for g in all_groups]
    all_groups = [g['group_id'] for g in all_groups]
    conn.close()
    return render_template('asprak.html', modules=modules, submissions=submissions, praktikans=praktikans,
                           courses=courses, sel_course=sel_course, admin=admin, aspraks=aspraks,
                           calculate_module_avg=calculate_module_avg, grade_legend=GRADE_LEGEND,
                           all_groups=all_groups, is_co_asprak=co_asprak, course_drive_link=course_drive_link,
                           active_tab=active_tab)

# ======== GRADE BATCH ========
@app.route('/asprak/grade_batch', methods=['POST'])
def asprak_grade_batch():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    gby = session['user_id']
    conn = get_db()
    for key, val in request.form.items():
        if key.startswith('praktikan_name_'):
            pid = key.split('_')[2]
            if val: conn.execute('UPDATE users SET name=? WHERE id=?', (val, pid))
        if key.startswith('pembukuan_'):
            pid = key.split('_')[1]
            conn.execute('UPDATE users SET pembukuan_score=? WHERE id=?', (val or 0, pid))
        if key.startswith('tp_score_'):
            parts = key.split('_'); pid, mid = parts[2], parts[3]
            tp = val or 0
            prak = request.form.get(f'praktikum_score_{pid}_{mid}', 0)
            modul = request.form.get(f'modul_score_{pid}_{mid}', 0)
            ex = conn.execute('SELECT id FROM grades WHERE praktikan_id=? AND module_id=?', (pid, mid)).fetchone()
            if ex:
                conn.execute('UPDATE grades SET tp_score=?, praktikum_score=?, modul_score=?, graded_by=? WHERE id=?',
                             (tp, prak, modul, gby, ex['id']))
            else:
                conn.execute('INSERT INTO grades (praktikan_id, module_id, tp_score, praktikum_score, modul_score, graded_by) VALUES (?,?,?,?,?,?)',
                             (pid, mid, tp, prak, modul, gby))
    conn.commit(); conn.close()
    flash('Semua perubahan berhasil disimpan!', 'success')
    course_id = request.form.get('course_id', '')
    return redirect(url_for('asprak_dashboard', course_id=course_id, tab=request.form.get('tab') or request.args.get('tab')))

# ======== EXPORT EXCEL ========
@app.route('/asprak/export')
def export_excel():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    sel_course = request.args.get('course_id', type=int)
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not sel_course: sel_course = user['course_id']
    course = conn.execute('SELECT name FROM courses WHERE id=?', (sel_course,)).fetchone()
    course_name = course['name'] if course else 'Unknown'
    modules = conn.execute('SELECT * FROM modules WHERE course_id=?', (sel_course,)).fetchall()
    ml = [dict(m) for m in modules]
    allowed = get_allowed_groups(user['name'], sel_course)
    if not allowed:
        flash('Tidak ada data', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    ph = ','.join('?' for _ in allowed)
    praks = conn.execute(f'SELECT * FROM users WHERE role="PRAKTIKAN" AND course_id=? AND group_id IN ({ph}) ORDER BY group_id, name',
                         (sel_course, *allowed)).fetchall()
    subs = conn.execute(f'SELECT module_id, group_id FROM submissions WHERE group_id IN ({ph})', allowed).fetchall()
    ss = {(s['module_id'], s['group_id']) for s in subs}
    data = []
    for p in praks:
        pd = dict(p)
        gr = conn.execute('SELECT * FROM grades WHERE praktikan_id=?', (p['id'],)).fetchall()
        pd['grades'] = {g['module_id']: dict(g) for g in gr}
        pd['total'] = calculate_total(pd['grades'], ml, pd.get('pembukuan_score', 0))
        pd['letter'] = get_letter_grade(pd['total'])
        data.append(pd)
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = f"Nilai {course_name}"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ct = Alignment(horizontal='center', vertical='center')
    headers = ['No', 'Nama', 'Kel']
    for m in ml:
        headers.extend([f'{m["name"]} TP', f'{m["name"]} Prak', f'{m["name"]} Modul', f'{m["name"]} Status'])
    headers.extend(['Pembukuan', 'Total', 'Huruf'])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hf; cell.fill = hfill; cell.alignment = ct; cell.border = tb
    for ri, p in enumerate(data, 2):
        ws.cell(row=ri, column=1, value=ri-1).border = tb
        ws.cell(row=ri, column=2, value=p['name']).border = tb
        ws.cell(row=ri, column=3, value=f"Kel {p['group_id']}").border = tb
        ci = 4
        for m in ml:
            g = p['grades'].get(m['id'], {})
            for v in [int(g.get('tp_score',0) or 0), int(g.get('praktikum_score',0) or 0), int(g.get('modul_score',0) or 0)]:
                ws.cell(row=ri, column=ci, value=v).border = tb; ci += 1
            st = '✅' if (m['id'], p['group_id']) in ss else '❌'
            ws.cell(row=ri, column=ci, value=st).border = tb; ci += 1
        ws.cell(row=ri, column=ci, value=int(p.get('pembukuan_score',0) or 0)).border = tb; ci += 1
        ws.cell(row=ri, column=ci, value=p['total']).border = tb; ci += 1
        ws.cell(row=ri, column=ci, value=p['letter']).border = tb
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(mx+3, 10)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Nilai_{course_name}_{user["name"]}.xlsx')

# ======== ADMIN: COURSE MANAGEMENT ========
@app.route('/admin/course/add', methods=['POST'])
def add_course():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    name = request.form.get('course_name', '').strip()
    desc = request.form.get('course_desc', '').strip()
    if name:
        conn = get_db()
        conn.execute('INSERT INTO courses (name, description) VALUES (?,?)', (name, desc))
        conn.commit(); conn.close()
        flash(f'Mata kuliah "{name}" berhasil ditambahkan!', 'success')
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

# ======== ADMIN: ASPRAK MANAGEMENT ========
@app.route('/admin/asprak/add', methods=['POST'])
def add_asprak():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    name = request.form.get('asprak_name', '').strip()
    pwd = request.form.get('asprak_password', '').strip()
    cid = request.form.get('asprak_course_id', type=int)
    is_co = 1 if request.form.get('is_co_asprak') else 0
    if name and pwd and cid:
        conn = get_db()
        ex = conn.execute('SELECT id FROM users WHERE LOWER(name)=LOWER(?) AND role="ASPRAK" AND course_id=?', (name, cid)).fetchone()
        if ex:
            flash(f'Asprak "{name}" sudah ada di MK ini', 'error')
        else:
            conn.execute('INSERT INTO users (name, role, group_id, password, course_id, is_admin, is_co_asprak) VALUES (?,?,?,?,?,?,?)',
                         (name, 'ASPRAK', 0, pwd, cid, 0, is_co))
            conn.commit()
            flash(f'Asprak "{name}" berhasil ditambahkan!', 'success')
        conn.close()
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/admin/asprak/edit', methods=['POST'])
def edit_asprak():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    aid = request.form.get('asprak_id', type=int)
    conn = get_db()
    target = conn.execute('SELECT * FROM users WHERE id=?', (aid,)).fetchone()
    if target and target['is_admin']:
        flash('Tidak bisa mengubah akun admin', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    new_name = request.form.get('edit_asprak_name', '').strip()
    new_pwd = request.form.get('edit_asprak_password', '').strip()
    if new_name: conn.execute('UPDATE users SET name=? WHERE id=?', (new_name, aid))
    if new_pwd: conn.execute('UPDATE users SET password=? WHERE id=?', (new_pwd, aid))
    conn.commit(); conn.close()
    flash('Asprak berhasil diperbarui!', 'success')
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/admin/asprak/delete', methods=['POST'])
def delete_asprak():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    aid = request.form.get('asprak_id', type=int)
    conn = get_db()
    target = conn.execute('SELECT * FROM users WHERE id=?', (aid,)).fetchone()
    if target and target['is_admin']:
        flash('Tidak bisa menghapus akun admin', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    conn.execute('DELETE FROM users WHERE id=?', (aid,))
    conn.commit(); conn.close()
    flash('Asprak berhasil dihapus', 'success')
    cid = request.form.get('course_id')
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

# ======== MODULE MANAGEMENT ========
@app.route('/asprak/module/toggle_status', methods=['POST'])
def toggle_module_status():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT is_co_asprak, is_admin FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user or not (user['is_co_asprak'] or user['is_admin']):
        flash('Hanya Co-Asprak atau Admin yang dapat membuka/tutup pengumpulan', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    mid = request.form.get('module_id'); cs = request.form.get('current_status', type=int)
    ns = 0 if cs == 1 else 1
    conn.execute('UPDATE modules SET is_open=? WHERE id=?', (ns, mid)); conn.commit(); conn.close()
    flash(f'Pengumpulan modul berhasil {"dibuka" if ns else "ditutup"}!', 'success')
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/module/add', methods=['POST'])
def add_module():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT is_co_asprak, is_admin FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user or not (user['is_co_asprak'] or user['is_admin']):
        flash('Hanya Co-Asprak atau Admin yang dapat menambah modul', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    name = request.form.get('name'); desc = request.form.get('description')
    deadline = request.form.get('deadline') or None
    cid = request.form.get('course_id', type=int)
    if name and cid:
        conn.execute('INSERT INTO modules (name, description, is_open, deadline, course_id) VALUES (?,?,1,?,?)', (name, desc, deadline, cid))
        conn.commit()
        flash(f'Modul {name} berhasil ditambahkan', 'success')
    conn.close()
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/module/edit', methods=['POST'])
def edit_module():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT is_co_asprak, is_admin FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user or not (user['is_co_asprak'] or user['is_admin']):
        flash('Hanya Co-Asprak atau Admin yang dapat mengedit modul', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    mid = request.form.get('module_id'); name = request.form.get('name'); desc = request.form.get('description')
    deadline = request.form.get('deadline') or None
    conn.execute('UPDATE modules SET name=?, description=?, deadline=? WHERE id=?', (name, desc, deadline, mid)); conn.commit(); conn.close()
    flash('Modul berhasil diperbarui', 'success')
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/module/delete', methods=['POST'])
def delete_module():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT is_co_asprak, is_admin FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user or not (user['is_co_asprak'] or user['is_admin']):
        flash('Hanya Co-Asprak atau Admin yang dapat menghapus modul', 'error'); conn.close(); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    mid = request.form.get('module_id')
    subs = conn.execute('SELECT * FROM submissions WHERE module_id=?', (mid,)).fetchall()
    for sub in subs:
        try: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sub['file_path']))
        except: pass
    conn.execute('DELETE FROM submissions WHERE module_id=?', (mid,))
    conn.execute('DELETE FROM grades WHERE module_id=?', (mid,))
    conn.execute('DELETE FROM modules WHERE id=?', (mid,))
    conn.commit(); conn.close()
    flash('Modul berhasil dihapus', 'success')
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/submissions/delete/<int:id>', methods=['POST'])
def delete_submission(id):
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db(); sub = conn.execute('SELECT * FROM submissions WHERE id=?', (id,)).fetchone()
    if sub:
        try: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sub['file_path']))
        except: pass
        conn.execute('DELETE FROM submissions WHERE id=?', (id,)); conn.commit()
        flash('Pengumpulan berhasil dihapus', 'success')
    conn.close()
    return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/uploads/<name>')
def download_file(name):
    return send_from_directory(app.config['UPLOAD_FOLDER'], name)

# ======== PRAKTIKAN MANAGEMENT (ALL ASPRAK) ========
@app.route('/asprak/praktikan/add', methods=['POST'])
def add_praktikan():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    name = request.form.get('new_praktikan_name', '').strip()
    gid = request.form.get('new_praktikan_group', type=int)
    cid = request.form.get('course_id', type=int)
    if name and gid and cid:
        conn = get_db()
        conn.execute('INSERT INTO users (name, role, group_id, password, course_id, is_admin, is_co_asprak, pembukuan_score) VALUES (?,?,?,?,?,?,?,?)',
                     (name, 'PRAKTIKAN', gid, None, cid, 0, 0, 0))
        conn.commit(); conn.close()
        flash(f'Praktikan "{name}" berhasil ditambahkan ke Kel {gid}!', 'success')
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/praktikan/edit_group', methods=['POST'])
def edit_praktikan_group():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    pid = request.form.get('praktikan_id', type=int)
    new_gid = request.form.get('new_group_id', type=int)
    cid = request.form.get('course_id', type=int)
    if pid and new_gid:
        conn = get_db()
        conn.execute('UPDATE users SET group_id=? WHERE id=?', (new_gid, pid))
        conn.commit(); conn.close()
        flash('Kelompok praktikan berhasil diubah!', 'success')
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

@app.route('/asprak/praktikan/delete', methods=['POST'])
def delete_praktikan():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    pid = request.form.get('praktikan_id', type=int)
    cid = request.form.get('course_id', type=int)
    if pid:
        conn = get_db()
        conn.execute('DELETE FROM users WHERE id=? AND role="PRAKTIKAN"', (pid,))
        conn.execute('DELETE FROM grades WHERE praktikan_id=?', (pid,))
        conn.commit()
        conn.close()
        flash('Praktikan berhasil dihapus', 'success')
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

# ======== ADMIN: CO-ASPRAK TOGGLE ========
@app.route('/admin/asprak/toggle_co', methods=['POST'])
def toggle_co_asprak():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    aid = request.form.get('asprak_id', type=int)
    conn = get_db()
    a = conn.execute('SELECT is_co_asprak FROM users WHERE id=?', (aid,)).fetchone()
    if a:
        nv = 0 if a['is_co_asprak'] else 1
        conn.execute('UPDATE users SET is_co_asprak=? WHERE id=?', (nv, aid))
        conn.commit()
        flash(f'Status Co-Asprak berhasil {"diaktifkan" if nv else "dinonaktifkan"}!', 'success')
    cid = request.form.get('course_id')
    conn.close()
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

# ======== ADMIN: ASSIGN ASPRAK TO COURSE ========
@app.route('/admin/asprak/assign_course', methods=['POST'])
def assign_asprak_course():
    if not is_admin_user(session.get('user_id')):
        flash('Akses ditolak', 'error'); return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    aid = request.form.get('asprak_id', type=int)
    cid = request.form.get('new_course_id', type=int)
    if aid and cid:
        conn = get_db()
        conn.execute('UPDATE users SET course_id=? WHERE id=?', (cid, aid))
        conn.commit(); conn.close()
        flash('Asprak berhasil dipindah ke praktikum baru!', 'success')
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))

# ======== CO-ASPRAK: SET DRIVE LINK ========
@app.route('/asprak/set_drive_link', methods=['POST'])
def set_drive_link():
    if 'role' not in session or session['role'] != 'ASPRAK':
        return redirect(url_for('asprak_login'))
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not (user['is_co_asprak'] or user['is_admin']):
        flash('Hanya Co-Asprak atau Admin yang dapat mengatur Drive folder', 'error')
        conn.close()
        return redirect(url_for('asprak_dashboard', tab=request.form.get('tab') or request.args.get('tab')))
    cid = request.form.get('course_id', type=int)
    link = request.form.get('drive_link', '').strip()
    if cid:
        conn.execute('UPDATE courses SET drive_link=? WHERE id=?', (link or None, cid))
        conn.commit()
        flash('Link Google Drive berhasil disimpan!' if link else 'Link Google Drive dihapus.', 'success')
    conn.close()
    return redirect(url_for('asprak_dashboard', course_id=cid, tab=request.form.get('tab') or request.args.get('tab')))


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=5000)
